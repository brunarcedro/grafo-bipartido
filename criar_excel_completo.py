# -*- coding: utf-8 -*-
# -*- coding: utf-8 -*-
"""
Cria arquivo Excel completo com as 4 planilhas do teste de mesa
Requer: pip install openpyxl
"""
"""
Autoras: Larissa Paganini e Bruna Cedro
Disciplina: Tópicos de Programação Avançada
Ano: 2024
"""



try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_DISPONIVEL = True
except ImportError:
    OPENPYXL_DISPONIVEL = False
    print("AVISO: openpyxl nao instalado. Instalando...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_DISPONIVEL = True


def criar_estilos():
    """Define estilos de formatação"""
    estilos = {
        'cabecalho': {
            'font': Font(name='Arial', size=11, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thick')
            )
        },
        'v1': {
            'fill': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
        },
        'v2': {
            'fill': PatternFill(start_color='4ECDC4', end_color='4ECDC4', fill_type='solid'),
        },
        'nao_visitado': {
            'fill': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'),
        },
        'normal': {
            'font': Font(name='Arial', size=10),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    }
    return estilos


def aplicar_estilo(celula, estilo_dict):
    """Aplica um estilo a uma célula"""
    if 'font' in estilo_dict:
        celula.font = estilo_dict['font']
    if 'fill' in estilo_dict:
        celula.fill = estilo_dict['fill']
    if 'alignment' in estilo_dict:
        celula.alignment = estilo_dict['alignment']
    if 'border' in estilo_dict:
        celula.border = estilo_dict['border']


def criar_planilha1_entrada(wb, estilos):
    """Cria Planilha 1: Grafo de Entrada"""
    ws = wb.create_sheet("1. Grafo de Entrada", 0)

    # Título
    ws['A1'] = 'TABELA 1: LISTA DE ARESTAS'
    ws['A1'].font = Font(name='Arial', size=12, bold=True)

    # Cabeçalho tabela 1
    headers1 = ['#', 'Usuário', 'Filme']
    for col, header in enumerate(headers1, 1):
        cell = ws.cell(row=2, column=col, value=header)
        aplicar_estilo(cell, estilos['cabecalho'])

    # Dados tabela 1
    dados1 = [
        (1, 'Alice', 'Matrix'),
        (2, 'Alice', 'Inception'),
        (3, 'Alice', 'Interstellar'),
        (4, 'Bob', 'Matrix'),
        (5, 'Bob', 'Avatar'),
        (6, 'Bob', 'Titanic'),
        (7, 'Carlos', 'Inception'),
        (8, 'Carlos', 'Avatar'),
        (9, 'Diana', 'Interstellar'),
        (10, 'Diana', 'Titanic'),
        (11, 'Eduardo', 'Matrix'),
        (12, 'Eduardo', 'Inception'),
    ]

    for row_idx, row_data in enumerate(dados1, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            aplicar_estilo(cell, estilos['normal'])

    # Tabela 2
    ws['A16'] = 'TABELA 2: LISTA DE ADJACÊNCIAS'
    ws['A16'].font = Font(name='Arial', size=12, bold=True)

    headers2 = ['Vértice', 'Tipo', 'Adjacentes']
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=17, column=col, value=header)
        aplicar_estilo(cell, estilos['cabecalho'])

    dados2 = [
        ('Alice', 'Usuário', 'Matrix, Inception, Interstellar'),
        ('Bob', 'Usuário', 'Matrix, Avatar, Titanic'),
        ('Carlos', 'Usuário', 'Inception, Avatar'),
        ('Diana', 'Usuário', 'Interstellar, Titanic'),
        ('Eduardo', 'Usuário', 'Matrix, Inception'),
        ('Matrix', 'Filme', 'Alice, Bob, Eduardo'),
        ('Inception', 'Filme', 'Alice, Carlos, Eduardo'),
        ('Interstellar', 'Filme', 'Alice, Diana'),
        ('Avatar', 'Filme', 'Bob, Carlos'),
        ('Titanic', 'Filme', 'Bob, Diana'),
    ]

    for row_idx, row_data in enumerate(dados2, 18):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            aplicar_estilo(cell, estilos['normal'])

    # Ajustar larguras
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35


def criar_planilha2_inicializacao(wb, estilos):
    """Cria Planilha 2: Inicialização"""
    ws = wb.create_sheet("2. Inicializacao", 1)

    ws['A1'] = 'ESTADO INICIAL DO ALGORITMO'
    ws['A1'].font = Font(name='Arial', size=12, bold=True)

    headers = ['Vértice', 'Cor Inicial', 'Tipo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        aplicar_estilo(cell, estilos['cabecalho'])

    dados = [
        ('Alice', 0, 'Usuário'),
        ('Bob', 0, 'Usuário'),
        ('Carlos', 0, 'Usuário'),
        ('Diana', 0, 'Usuário'),
        ('Eduardo', 0, 'Usuário'),
        ('Matrix', 0, 'Filme'),
        ('Inception', 0, 'Filme'),
        ('Interstellar', 0, 'Filme'),
        ('Avatar', 0, 'Filme'),
        ('Titanic', 0, 'Filme'),
    ]

    for row_idx, row_data in enumerate(dados, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            aplicar_estilo(cell, estilos['normal'])
            if col_idx == 2:  # Coluna de cor
                aplicar_estilo(cell, estilos['nao_visitado'])

    # Legenda
    ws['A15'] = 'LEGENDA'
    ws['A15'].font = Font(name='Arial', size=12, bold=True)

    headers_leg = ['Código', 'Significado', 'Cor']
    for col, header in enumerate(headers_leg, 1):
        cell = ws.cell(row=16, column=col, value=header)
        aplicar_estilo(cell, estilos['cabecalho'])

    ws.cell(row=17, column=1, value=0)
    ws.cell(row=17, column=2, value='Não visitado')
    aplicar_estilo(ws.cell(row=17, column=3), estilos['nao_visitado'])

    ws.cell(row=18, column=1, value=1)
    ws.cell(row=18, column=2, value='Conjunto V1')
    aplicar_estilo(ws.cell(row=18, column=3), estilos['v1'])

    ws.cell(row=19, column=1, value=2)
    ws.cell(row=19, column=2, value='Conjunto V2')
    aplicar_estilo(ws.cell(row=19, column=3), estilos['v2'])

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15


def criar_planilha3_execucao(wb, estilos):
    """Cria Planilha 3: Execução BFS"""
    ws = wb.create_sheet("3. Execucao BFS", 2)

    ws['A1'] = 'EXECUÇÃO DO ALGORITMO BFS (26 PASSOS)'
    ws['A1'].font = Font(name='Arial', size=12, bold=True)

    headers = ['Passo', 'Ação', 'Vértice Atual', 'Vértice Visitado', 'Cor', 'Fila', 'Observação']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        aplicar_estilo(cell, estilos['cabecalho'])

    # Dados simplificados (primeiros passos importantes)
    dados = [
        (0, 'Inicializar', '-', '-', '-', '[]', 'Todos não visitados'),
        (1, 'Iniciar BFS', 'Alice', 'Alice', 1, '[Alice]', 'Primeiro vértice'),
        (2, 'Processar', 'Alice', 'Matrix', 2, '[Matrix]', 'Adjacente de Alice'),
        (3, 'Processar', 'Alice', 'Inception', 2, '[Matrix, Inception]', 'Adjacente de Alice'),
        (4, 'Processar', 'Alice', 'Interstellar', 2, '[Matrix, Inception, Interstellar]', 'Adjacente de Alice'),
        (5, 'Processar', 'Matrix', 'Alice', '-', '[Inception, Interstellar]', 'Já visitado OK'),
        (6, 'Processar', 'Matrix', 'Bob', 1, '[Inception, Interstellar, Bob]', 'Adjacente de Matrix'),
        (7, 'Processar', 'Matrix', 'Eduardo', 1, '[Inception, Interstellar, Bob, Eduardo]', 'Adjacente de Matrix'),
        (8, 'Processar', 'Inception', 'Alice', '-', '[Interstellar, Bob, Eduardo]', 'Já visitado OK'),
        (9, 'Processar', 'Inception', 'Carlos', 1, '[Interstellar, Bob, Eduardo, Carlos]', 'Adjacente de Inception'),
        (10, 'Processar', 'Inception', 'Eduardo', '-', '[Interstellar, Bob, Eduardo, Carlos]', 'Já visitado OK'),
        (11, 'Processar', 'Interstellar', 'Alice', '-', '[Bob, Eduardo, Carlos]', 'Já visitado OK'),
        (12, 'Processar', 'Interstellar', 'Diana', 1, '[Bob, Eduardo, Carlos, Diana]', 'Adjacente de Interstellar'),
        (13, 'Processar', 'Bob', 'Matrix', '-', '[Eduardo, Carlos, Diana]', 'Já visitado OK'),
        (14, 'Processar', 'Bob', 'Avatar', 2, '[Eduardo, Carlos, Diana, Avatar]', 'Adjacente de Bob'),
        (15, 'Processar', 'Bob', 'Titanic', 2, '[Eduardo, Carlos, Diana, Avatar, Titanic]', 'Adjacente de Bob'),
        (16, 'Processar', 'Eduardo', 'Matrix', '-', '[Carlos, Diana, Avatar, Titanic]', 'Já visitado OK'),
        (17, 'Processar', 'Eduardo', 'Inception', '-', '[Carlos, Diana, Avatar, Titanic]', 'Já visitado OK'),
        (18, 'Processar', 'Carlos', 'Inception', '-', '[Diana, Avatar, Titanic]', 'Já visitado OK'),
        (19, 'Processar', 'Carlos', 'Avatar', '-', '[Diana, Avatar, Titanic]', 'Já visitado OK'),
        (20, 'Processar', 'Diana', 'Interstellar', '-', '[Avatar, Titanic]', 'Já visitado OK'),
        (21, 'Processar', 'Diana', 'Titanic', '-', '[Avatar, Titanic]', 'Já visitado OK'),
        (22, 'Processar', 'Avatar', 'Bob', '-', '[Titanic]', 'Já visitado OK'),
        (23, 'Processar', 'Avatar', 'Carlos', '-', '[Titanic]', 'Já visitado OK'),
        (24, 'Processar', 'Titanic', 'Bob', '-', '[]', 'Já visitado OK'),
        (25, 'Processar', 'Titanic', 'Diana', '-', '[]', 'Já visitado OK'),
        (26, 'Finalizar', '-', '-', '-', '[]', 'ALGORITMO COMPLETO'),
    ]

    for row_idx, row_data in enumerate(dados, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            aplicar_estilo(cell, estilos['normal'])

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 25


def criar_planilha4_resultado(wb, estilos):
    """Cria Planilha 4: Resultado"""
    ws = wb.create_sheet("4. Resultado", 3)

    ws['A1'] = 'RESULTADO: O GRAFO É BIPARTIDO'
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color='008000')

    # Partição
    ws['A3'] = 'PARTIÇÃO BIPARTIDA'
    ws['A3'].font = Font(name='Arial', size=12, bold=True)

    headers1 = ['Conjunto', 'Cor', 'Vértices', 'Quantidade']
    for col, header in enumerate(headers1, 1):
        cell = ws.cell(row=4, column=col, value=header)
        aplicar_estilo(cell, estilos['cabecalho'])

    ws.cell(row=5, column=1, value='V1')
    ws.cell(row=5, column=2, value='1 (Vermelho)')
    ws.cell(row=5, column=3, value='Alice, Bob, Carlos, Diana, Eduardo')
    ws.cell(row=5, column=4, value=5)
    aplicar_estilo(ws.cell(row=5, column=2), estilos['v1'])

    ws.cell(row=6, column=1, value='V2')
    ws.cell(row=6, column=2, value='2 (Azul)')
    ws.cell(row=6, column=3, value='Matrix, Inception, Interstellar, Avatar, Titanic')
    ws.cell(row=6, column=4, value=5)
    aplicar_estilo(ws.cell(row=6, column=2), estilos['v2'])

    # Estatísticas
    ws['A9'] = 'ESTATÍSTICAS'
    ws['A9'].font = Font(name='Arial', size=12, bold=True)

    stats = [
        ('Total de vértices', 10),
        ('Total de arestas', 12),
        ('Vértices em V1', 5),
        ('Vértices em V2', 5),
        ('É bipartido?', 'SIM'),
        ('Componentes conexos', 1),
    ]

    for row_idx, (metrica, valor) in enumerate(stats, 10):
        ws.cell(row=row_idx, column=1, value=metrica)
        ws.cell(row=row_idx, column=2, value=valor)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 12


def main():
    """Função principal"""
    print("="*60)
    print("CRIANDO ARQUIVO EXCEL COMPLETO")
    print("="*60)

    # Criar workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove planilha padrão

    # Criar estilos
    estilos = criar_estilos()

    # Criar planilhas
    print("\n1. Criando planilha: Grafo de Entrada...")
    criar_planilha1_entrada(wb, estilos)

    print("2. Criando planilha: Inicializacao...")
    criar_planilha2_inicializacao(wb, estilos)

    print("3. Criando planilha: Execucao BFS...")
    criar_planilha3_execucao(wb, estilos)

    print("4. Criando planilha: Resultado...")
    criar_planilha4_resultado(wb, estilos)

    # Salvar
    arquivo = 'teste_de_mesa_completo.xlsx'
    print(f"\n5. Salvando arquivo: {arquivo}...")
    wb.save(arquivo)

    print("\n" + "="*60)
    print("SUCESSO! Arquivo Excel criado!")
    print("="*60)
    print(f"\nArquivo: {arquivo}")
    print("\nConteudo:")
    print("  - Aba 1: Grafo de Entrada (arestas e adjacencias)")
    print("  - Aba 2: Inicializacao (estado inicial)")
    print("  - Aba 3: Execucao BFS (26 passos)")
    print("  - Aba 4: Resultado (particao e estatisticas)")
    print("\nAbra o arquivo no Excel/LibreOffice!")
    print("="*60)


if __name__ == "__main__":
    main()
